import io
from collections import defaultdict

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from Backend.configurations import (
    BUS_CONFIG,
    ROUTES_CONFIG,
    STATIONS_CONFIG,
)


TIME_SLOT_MINUTES = 5

HEADER_FILL = "1F1F1F"
HEADER_FONT = "FFFFFF"

TRAVEL_FILL = "9DC3E6"
WAIT_FILL = "F4B183"
CHARGING_FILL = "A9D18E"
ARRIVED_FILL = "D9D9D9"
ISSUE_FILL = "FF6666"
IDLE_FILL = "FFFFFF"
SUMMARY_FILL = "FFF2CC"

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def build_excel_report(result) -> bytes:
    """Create the downloadable Excel workbook.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(_timeline_summary_rows(result)).to_excel(
            writer,
            index=False,
            sheet_name="Timeline Summary",
        )

        _bus_timeline_km_dataframe(result).to_excel(
            writer,
            index=False,
            sheet_name="Bus Timeline Km",
        )

        _charger_timeline_dataframe(result).to_excel(
            writer,
            index=False,
            sheet_name="Charger Timeline View",
        )

        pd.DataFrame(_summary_rows(result)).to_excel(
            writer,
            index=False,
            sheet_name="Summary",
        )

        pd.DataFrame(_operator_metric_rows(result)).to_excel(
            writer,
            index=False,
            sheet_name="Operator Metrics",
        )

        pd.DataFrame(_bus_timetable_rows(result)).to_excel(
            writer,
            index=False,
            sheet_name="Bus Timetable",
        )

        pd.DataFrame(_bus_charging_event_rows(result)).to_excel(
            writer,
            index=False,
            sheet_name="Bus Charging Events",
        )

        pd.DataFrame(_station_order_rows(result)).to_excel(
            writer,
            index=False,
            sheet_name="Station Orders",
        )

        _style_workbook(writer)

    output.seek(0)
    return output


def _summary_rows(result) -> object:
    """Build summary metrics from the schedule.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    rows = [
        {"Metric": "Scenario ID", "Value": result.get("scenario_id")},
        {"Metric": "Scenario Name", "Value": result.get("scenario_name")},
        {"Metric": "Status", "Value": result.get("status")},
        {"Metric": "Objective Value", "Value": result.get("objective_value")},
    ]

    for key, value in result.get("summary", {}).items():
        if not isinstance(value, dict):
            rows.append({"Metric": key, "Value": value})

    for key, value in result.get("optimization_weights", {}).items():
        rows.append({"Metric": f"weight_{key}", "Value": value})

    return rows


def _operator_metric_rows(result) -> object:
    """Handle operator metric rows logic.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    summary = result.get("summary", {})

    operator_ids = sorted(
        set(summary.get("operator_average_wait_minutes", {}))
        | set(summary.get("operator_total_wait_minutes", {}))
        | set(summary.get("operator_bus_count", {}))
    )

    return [
        {
            "Operator": operator_id,
            "Bus Count": summary.get("operator_bus_count", {}).get(operator_id, 0),
            "Total Wait Minutes": summary.get("operator_total_wait_minutes", {}).get(operator_id, 0),
            "Average Wait Minutes": summary.get("operator_average_wait_minutes", {}).get(operator_id, 0),
            "Max Wait Minutes": summary.get("operator_max_wait_minutes", {}).get(operator_id, 0),
        }
        for operator_id in operator_ids
    ]


def _bus_timetable_rows(result) -> object:
    """Convert or compare schedule time values.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    return [
        {
            "Bus ID": bus["bus_id"],
            "Operator": bus["operator_id"],
            "Route": bus["route_id"],
            "Origin": bus["origin"],
            "Destination": bus["destination"],
            "Departure Time": bus["departure_time"],
            "Charging Plan": " -> ".join(bus["charging_plan"]),
            "Total Wait Minutes": bus["total_wait_minutes"],
            "Total Charging Stops": bus["total_charging_stops"],
            "Final Arrival Time": bus["final_arrival_time"],
        }
        for bus in result.get("bus_timetables", [])
    ]


def _bus_charging_event_rows(result) -> object:
    """Handle timeline event queue processing.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    rows = []

    for bus in result.get("bus_timetables", []):
        for event_number, event in enumerate(bus.get("charging_events", []), start=1):
            rows.append({
                "Bus ID": bus["bus_id"],
                "Operator": bus["operator_id"],
                "Route": bus["route_id"],
                "Event Number": event_number,
                "Station ID": event["station_id"],
                "Charger ID": event["charger_id"],
                "Reached At": event["reached_at"],
                "Charging Started At": event["started_at"],
                "Charging Ended At": event["ended_at"],
                "Wait Minutes": event["wait_minutes"],
                "Charging Mode": event.get("charging_mode"),
                "Charger Blocked Minutes": event.get("charger_blocked_minutes"),
                "Operational Failure ID": event.get("operational_failure_id"),
                "Operational Failure Reason": event.get("operational_failure_reason"),
            })

    return rows


def _station_order_rows(result) -> object:
    """Build or validate station-level scheduling data.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    rows = []

    for station_id, events in result.get("station_charging_orders", {}).items():
        for order_number, event in enumerate(events, start=1):
            rows.append({
                "Station ID": station_id,
                "Order": order_number,
                "Bus ID": event["bus_id"],
                "Operator": event["operator_id"],
                "Charger ID": event["charger_id"],
                "Charging Started At": event["charging_started_at"],
                "Charging Ended At": event["charging_ended_at"],
                "Wait Minutes": event["wait_minutes"],
                "Charging Mode": event.get("charging_mode"),
                "Charger Blocked Minutes": event.get("charger_blocked_minutes"),
                "Operational Failure ID": event.get("operational_failure_id"),
                "Operational Failure Reason": event.get("operational_failure_reason"),
            })

    return rows


def _timeline_summary_rows(result) -> list:
    """Convert or compare schedule time values.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    bus_timeline_data = _bus_timeline_data(result)
    charger_events = _charger_events(result)

    completed_buses = sum(
        1
        for bus_data in bus_timeline_data
        if bus_data["completed"] == "YES"
    )

    overlap_count = sum(
        1
        for event in charger_events
        if event["overlap"] == "YES"
    )

    return [
        {
            "Check": "Scenario ID",
            "Result": result.get("scenario_id"),
            "Notes": "",
        },
        {
            "Check": "Scenario Name",
            "Result": result.get("scenario_name"),
            "Notes": "",
        },
        {
            "Check": "Solver Status",
            "Result": result.get("status"),
            "Notes": "",
        },
        {
            "Check": "Total Buses",
            "Result": len(bus_timeline_data),
            "Notes": "",
        },
        {
            "Check": "Completed Buses",
            "Result": completed_buses,
            "Notes": "Completed means distance covered equals total route distance.",
        },
        {
            "Check": "Incomplete Buses",
            "Result": len(bus_timeline_data) - completed_buses,
            "Notes": "Should be 0 for a valid schedule.",
        },
        {
            "Check": "Any Charger Overlap",
            "Result": "YES" if overlap_count else "NO",
            "Notes": "Should be NO. Same charger should not be used by overlapping buses.",
        },
        {
            "Check": "Battery Range Limit",
            "Result": BUS_CONFIG["maximum_range_km"],
            "Notes": "Distance since last full charge should not exceed this value.",
        },
        {
            "Check": "Timeline Slot Size",
            "Result": f"{TIME_SLOT_MINUTES} minutes",
            "Notes": "Each timeline column represents this time interval.",
        },
        {
            "Check": "Legend",
            "Result": "T = Travel, W = Wait, C = Charging, ARR = Arrived",
            "Notes": "Bus Timeline Km shows distance since last full charge inside each cell.",
        },
    ]


def _bus_timeline_km_dataframe(result) -> object:
    """Convert or compare schedule time values.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    bus_timeline_data = _bus_timeline_data(result)
    start_minute, end_minute = _timeline_bounds_from_bus_data(bus_timeline_data)
    time_slots = list(range(start_minute, end_minute + 1, TIME_SLOT_MINUTES))

    rows = []

    for bus_data in bus_timeline_data:
        row = {
            "Bus ID": bus_data["bus_id"],
            "Operator": bus_data["operator_id"],
            "Route": bus_data["route_id"],
            "Departure": bus_data["departure_time"],
            "Final Arrival": bus_data["final_arrival_time"],
            "Route Distance km": bus_data["route_distance_km"],
            "Distance Covered km": bus_data["distance_covered_km"],
            "Completed?": bus_data["completed"],
            "Max Km Between Charges": bus_data["max_km_between_charges"],
            "Range Valid?": bus_data["range_valid"],
            "Charging Plan": " -> ".join(bus_data["charging_plan"]),
        }

        for slot_start in time_slots:
            row[_minutes_to_time(slot_start)] = _bus_timeline_cell(
                bus_data,
                slot_start,
                slot_start + TIME_SLOT_MINUTES,
            )

        rows.append(row)

    return pd.DataFrame(rows)


def _charger_timeline_dataframe(result) -> object:
    """Convert or compare schedule time values.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    charger_events = _charger_events(result)

    if not charger_events:
        return pd.DataFrame()

    start_minute = min(event["start_minute"] for event in charger_events)
    end_minute = max(event["end_minute"] for event in charger_events)
    time_slots = list(range(start_minute, end_minute + 1, TIME_SLOT_MINUTES))

    charger_ids = sorted(
        charger_id
        for station in STATIONS_CONFIG
        for charger_id in [
            f"{station['station_id']}-{charger_number}"
            for charger_number in range(1, station["charger_count"] + 1)
        ]
    )

    events_by_charger = defaultdict(list)

    for event in charger_events:
        events_by_charger[event["charger_id"]].append(event)

    rows = []

    for charger_id in charger_ids:
        station_id = charger_id.split("-")[0]

        row = {
            "Station": station_id,
            "Charger": charger_id,
            "Overlap Check": "YES"
            if any(event["overlap"] == "YES" for event in events_by_charger.get(charger_id, []))
            else "NO",
        }

        for slot_start in time_slots:
            row[_minutes_to_time(slot_start)] = _charger_timeline_cell(
                events_by_charger.get(charger_id, []),
                slot_start,
                slot_start + TIME_SLOT_MINUTES,
            )

        rows.append(row)

    return pd.DataFrame(rows)


def _bus_timeline_data(result) -> object:
    """Convert or compare schedule time values.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    routes = {
        route["route_id"]: route
        for route in ROUTES_CONFIG
    }

    bus_data_rows = []

    for bus in result.get("bus_timetables", []):
        route = routes[bus["route_id"]]
        route_distance = _route_distance(route)
        departure_minute = _time_to_minutes(bus["departure_time"])
        final_arrival_minute = _time_to_minutes_after(
            bus["final_arrival_time"],
            departure_minute,
        )

        activities = []
        current_station = bus["origin"]
        current_minute = departure_minute
        distance_since_charge = 0
        distance_covered = 0
        max_km_between_charges = 0

        for event in bus.get("charging_events", []):
            station_id = event["station_id"]

            reached_minute = _time_to_minutes_after(
                event["reached_at"],
                current_minute,
            )

            charging_started_minute = _time_to_minutes_after(
                event["started_at"],
                reached_minute,
            )

            charging_ended_minute = _time_to_minutes_after(
                event["ended_at"],
                charging_started_minute,
            )

            travel_distance = _distance_between(
                route,
                current_station,
                station_id,
            )

            travel_start_km = distance_since_charge
            travel_end_km = distance_since_charge + travel_distance

            activities.append({
                "activity": "TRAVEL",
                "from": current_station,
                "to": station_id,
                "station": station_id,
                "charger": "",
                "start_minute": current_minute,
                "end_minute": reached_minute,
                "start_km_since_charge": travel_start_km,
                "end_km_since_charge": travel_end_km,
                "distance_km": travel_distance,
            })

            distance_since_charge = travel_end_km
            distance_covered += travel_distance
            max_km_between_charges = max(max_km_between_charges, distance_since_charge)

            if charging_started_minute > reached_minute:
                activities.append({
                    "activity": "WAIT",
                    "from": "",
                    "to": station_id,
                    "station": station_id,
                    "charger": "",
                    "start_minute": reached_minute,
                    "end_minute": charging_started_minute,
                    "start_km_since_charge": distance_since_charge,
                    "end_km_since_charge": distance_since_charge,
                    "distance_km": 0,
                })

            activities.append({
                "activity": "CHARGING",
                "from": "",
                "to": station_id,
                "station": station_id,
                "charger": event["charger_id"],
                "start_minute": charging_started_minute,
                "end_minute": charging_ended_minute,
                "start_km_since_charge": distance_since_charge,
                "end_km_since_charge": 0,
                "distance_km": 0,
            })

            current_station = station_id
            current_minute = charging_ended_minute
            distance_since_charge = 0

        final_travel_distance = _distance_between(
            route,
            current_station,
            bus["destination"],
        )

        activities.append({
            "activity": "TRAVEL",
            "from": current_station,
            "to": bus["destination"],
            "station": bus["destination"],
            "charger": "",
            "start_minute": current_minute,
            "end_minute": final_arrival_minute,
            "start_km_since_charge": distance_since_charge,
            "end_km_since_charge": distance_since_charge + final_travel_distance,
            "distance_km": final_travel_distance,
        })

        distance_since_charge += final_travel_distance
        distance_covered += final_travel_distance
        max_km_between_charges = max(max_km_between_charges, distance_since_charge)

        completed = "YES" if round(distance_covered, 2) == round(route_distance, 2) else "NO"

        bus_data_rows.append({
            "bus_id": bus["bus_id"],
            "operator_id": bus["operator_id"],
            "route_id": bus["route_id"],
            "origin": bus["origin"],
            "destination": bus["destination"],
            "departure_time": bus["departure_time"],
            "final_arrival_time": bus["final_arrival_time"],
            "charging_plan": bus["charging_plan"],
            "route_distance_km": route_distance,
            "distance_covered_km": round(distance_covered, 2),
            "completed": completed,
            "max_km_between_charges": round(max_km_between_charges, 2),
            "range_valid": "YES"
            if max_km_between_charges <= BUS_CONFIG["maximum_range_km"]
            else "NO",
            "activities": activities,
            "departure_minute": departure_minute,
            "final_arrival_minute": final_arrival_minute,
        })

    return bus_data_rows


def _charger_events(result) -> object:
    """Build or validate charger availability and assignment data.
    
    Args:
        result (dict): Final scheduler result dictionary.
    """
    events = []

    for station_id, station_events in result.get("station_charging_orders", {}).items():
        for event in station_events:
            start_minute = _time_to_minutes(event["charging_started_at"])
            end_minute = _time_to_minutes_after(
                event["charging_ended_at"],
                start_minute,
            )

            events.append({
                "station_id": station_id,
                "charger_id": event["charger_id"],
                "bus_id": event["bus_id"],
                "operator_id": event["operator_id"],
                "start_minute": start_minute,
                "end_minute": end_minute,
                "charging_mode": event.get("charging_mode", "NORMAL"),
                "overlap": "NO",
            })

    events_by_charger = defaultdict(list)

    for event in events:
        events_by_charger[event["charger_id"]].append(event)

    for charger_event_list in events_by_charger.values():
        sorted_events = sorted(
            charger_event_list,
            key=lambda item: item["start_minute"],
        )

        for index in range(1, len(sorted_events)):
            previous_event = sorted_events[index - 1]
            current_event = sorted_events[index]

            if current_event["start_minute"] < previous_event["end_minute"]:
                previous_event["overlap"] = "YES"
                current_event["overlap"] = "YES"

    return events


def _bus_timeline_cell(bus_data, slot_start, slot_end) -> str:
    """Convert or compare schedule time values.
    
    Args:
        bus_data (_type_): Bus data used by this function.
        slot_start (_type_): Slot start used by this function.
        slot_end (_type_): Slot end used by this function.
    """
    if slot_start >= bus_data["final_arrival_minute"]:
        return f"ARR\n{bus_data['route_distance_km']}km total"

    for activity in bus_data["activities"]:
        if slot_start < activity["end_minute"] and slot_end > activity["start_minute"]:
            if activity["activity"] == "TRAVEL":
                start_km, end_km = _km_range_for_slot(
                    activity,
                    slot_start,
                    slot_end,
                )

                return (
                    f"T:{activity['from']}->{activity['to']}\n"
                    f"{start_km}-{end_km}km"
                )

            if activity["activity"] == "WAIT":
                return (
                    f"W:{activity['station']}\n"
                    f"{round(activity['start_km_since_charge'], 1)}km"
                )

            if activity["activity"] == "CHARGING":
                return (
                    f"C:{activity['station']}/{activity['charger']}\n"
                    f"RESET→0km"
                )

    return ""


def _charger_timeline_cell(events, slot_start, slot_end) -> object:
    """Convert or compare schedule time values.
    
    Args:
        events (_type_): Events used by this function.
        slot_start (_type_): Slot start used by this function.
        slot_end (_type_): Slot end used by this function.
    """
    overlapping_events = [
        event
        for event in events
        if slot_start < event["end_minute"] and slot_end > event["start_minute"]
    ]

    if not overlapping_events:
        return ""

    if len(overlapping_events) > 1:
        return "OVERLAP"

    return overlapping_events[0]["bus_id"]


def _km_range_for_slot(activity, slot_start, slot_end) -> tuple:
    """Handle km range for slot logic.
    
    Args:
        activity (_type_): Activity used by this function.
        slot_start (_type_): Slot start used by this function.
        slot_end (_type_): Slot end used by this function.
    """
    activity_start = activity["start_minute"]
    activity_end = activity["end_minute"]
    duration = max(activity_end - activity_start, 1)

    clipped_start = max(slot_start, activity_start)
    clipped_end = min(slot_end, activity_end)

    start_ratio = (clipped_start - activity_start) / duration
    end_ratio = (clipped_end - activity_start) / duration

    start_km = activity["start_km_since_charge"] + activity["distance_km"] * start_ratio
    end_km = activity["start_km_since_charge"] + activity["distance_km"] * end_ratio

    return round(start_km, 1), round(end_km, 1)


def _timeline_bounds_from_bus_data(bus_timeline_data) -> tuple:
    """Convert or compare schedule time values.
    
    Args:
        bus_timeline_data (_type_): Bus timeline data used by this function.
    """
    if not bus_timeline_data:
        return 0, 0

    start_minute = min(
        bus_data["departure_minute"]
        for bus_data in bus_timeline_data
    )

    end_minute = max(
        bus_data["final_arrival_minute"]
        for bus_data in bus_timeline_data
    )

    return _floor_to_slot(start_minute), _ceil_to_slot(end_minute)


def _style_workbook(writer) -> None:
    """Handle style workbook logic.
    
    Args:
        writer (_type_): Writer used by this function.
    """
    for worksheet in writer.book.worksheets:
        _style_header_row(worksheet)
        _autosize_columns(worksheet)
        _freeze_header_only(worksheet)

    if "Timeline Summary" in writer.sheets:
        _style_timeline_summary(writer.sheets["Timeline Summary"])

    if "Bus Timeline Km" in writer.sheets:
        _style_bus_timeline_sheet(writer.sheets["Bus Timeline Km"])

    if "Charger Timeline View" in writer.sheets:
        _style_charger_timeline_sheet(writer.sheets["Charger Timeline View"])


def _style_header_row(worksheet) -> None:
    """Handle style header row logic.
    
    Args:
        worksheet (_type_): OpenPyXL worksheet being written.
    """
    for cell in worksheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _style_timeline_summary(worksheet) -> None:
    """Convert or compare schedule time values.
    
    Args:
        worksheet (_type_): OpenPyXL worksheet being written.
    """
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=SUMMARY_FILL)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER


def _style_bus_timeline_sheet(worksheet) -> None:
    """Convert or compare schedule time values.
    
    Args:
        worksheet (_type_): OpenPyXL worksheet being written.
    """
    fixed_columns = 11

    worksheet.freeze_panes = "A2"

    for row in worksheet.iter_rows(min_row=2):
        completed_cell = row[7]
        range_valid_cell = row[9]

        if completed_cell.value != "YES" or range_valid_cell.value != "YES":
            for cell in row[:fixed_columns]:
                cell.fill = PatternFill("solid", fgColor=ISSUE_FILL)

        for cell in row[fixed_columns:]:
            value = str(cell.value or "")

            if value.startswith("T:"):
                fill = TRAVEL_FILL
            elif value.startswith("W:"):
                fill = WAIT_FILL
            elif value.startswith("C:"):
                fill = CHARGING_FILL
            elif value.startswith("ARR"):
                fill = ARRIVED_FILL
            elif value:
                fill = ISSUE_FILL
            else:
                fill = IDLE_FILL

            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 34

    for column_number in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = (
            14 if column_number <= fixed_columns else 16
        )


def _style_charger_timeline_sheet(worksheet) -> None:
    """Convert or compare schedule time values.
    
    Args:
        worksheet (_type_): OpenPyXL worksheet being written.
    """
    fixed_columns = 3

    worksheet.freeze_panes = "A2"

    for row in worksheet.iter_rows(min_row=2):
        overlap_check = row[2].value

        if overlap_check == "YES":
            for cell in row[:fixed_columns]:
                cell.fill = PatternFill("solid", fgColor=ISSUE_FILL)

        for cell in row[fixed_columns:]:
            value = str(cell.value or "")

            if value == "OVERLAP":
                fill = ISSUE_FILL
            elif value:
                fill = CHARGING_FILL
            else:
                fill = IDLE_FILL

            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    for row_number in range(2, worksheet.max_row + 1):
        worksheet.row_dimensions[row_number].height = 28

    for column_number in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_number)].width = (
            14 if column_number <= fixed_columns else 13
        )


def _autosize_columns(worksheet) -> None:
    """Handle autosize columns logic.
    
    Args:
        worksheet (_type_): OpenPyXL worksheet being written.
    """
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 10),
            32,
        )


def _freeze_header_only(worksheet) -> None:
    """Handle freeze header only logic.
    
    Args:
        worksheet (_type_): OpenPyXL worksheet being written.
    """
    worksheet.freeze_panes = "A2"


def _route_distance(route) -> object:
    """Calculate route, station, or distance information.
    
    Args:
        route (_type_): Route used by this function.
    """
    return sum(
        segment["distance_km"]
        for segment in route["station_distances_in_km"]
    )


def _distance_between(route, from_station, to_station) -> object:
    """Calculate distance-related values.
    
    Args:
        route (_type_): Route used by this function.
        from_station (_type_): From station used by this function.
        to_station (_type_): To station used by this function.
    """
    cumulative = _cumulative_distance(route)
    return abs(cumulative[to_station] - cumulative[from_station])


def _cumulative_distance(route) -> object:
    """Calculate distance-related values.
    
    Args:
        route (_type_): Route used by this function.
    """
    total_distance = 0
    cumulative = {route["station_sequence"][0]: 0}

    for segment in route["station_distances_in_km"]:
        total_distance += segment["distance_km"]
        cumulative[segment["to_station"]] = total_distance

    return cumulative


def _time_to_minutes(time_text) -> object:
    """Convert or compare schedule time values.
    
    Args:
        time_text (str): Time value in HH:MM format.
    """
    hour, minute = map(int, time_text.split(":"))
    return hour * 60 + minute


def _time_to_minutes_after(time_text, minimum_minute) -> object:
    """Convert or compare schedule time values.
    
    Args:
        time_text (str): Time value in HH:MM format.
        minimum_minute (_type_): Minimum minute represented in minutes.
    """
    candidate_minute = _time_to_minutes(time_text)

    while candidate_minute < minimum_minute:
        candidate_minute += 24 * 60

    return candidate_minute


def _minutes_to_time(minutes) -> str:
    """Convert or compare schedule time values.
    
    Args:
        minutes (_type_): Minutes used by this function.
    """
    hour = (minutes // 60) % 24
    minute = minutes % 60

    return f"{hour:02d}:{minute:02d}"


def _floor_to_slot(minute) -> object:
    """Handle floor to slot logic.
    
    Args:
        minute (int): Timeline minute used for ordering events.
    """
    return (minute // TIME_SLOT_MINUTES) * TIME_SLOT_MINUTES


def _ceil_to_slot(minute) -> object:
    """Handle ceil to slot logic.
    
    Args:
        minute (int): Timeline minute used for ordering events.
    """
    return ((minute + TIME_SLOT_MINUTES - 1) // TIME_SLOT_MINUTES) * TIME_SLOT_MINUTES
